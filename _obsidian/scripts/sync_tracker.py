from __future__ import annotations

import csv
import math
import re
import warnings
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "25_application_tracker_template.xlsx"
GENERATED = ROOT / "_obsidian" / "generated"
CSV_DIR = GENERATED / "csv"
MD_DIR = GENERATED / "md"

SYNCED_SHEETS = [
    "Applications",
    "Companies",
    "Contacts",
    "Interviews",
    "Offers",
    "Weekly Metrics",
    "Skill Sprint",
    "Content Assets",
]


def slug(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return int(value)
        return round(value, 4)
    return value


def display(value: Any) -> str:
    value = clean_value(value)
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def md_escape(value: Any) -> str:
    return display(value).replace("|", "\\|")


def find_table_ref(ws: Any) -> str | None:
    if not ws.tables:
        return None
    first_name = next(iter(ws.tables.keys()))
    return ws.tables[first_name].ref


def extract_table(ws_values: Any, ws_formulas: Any) -> tuple[list[str], list[dict[str, Any]]]:
    ref = find_table_ref(ws_formulas)
    if ref:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws_formulas.max_column or 1, ws_formulas.max_row or 1

    headers = []
    for cell in ws_formulas.iter_rows(
        min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True
    ):
        headers = [display(v) for v in cell]
        break

    # Ignore trailing empty header columns.
    while headers and not headers[-1]:
        headers.pop()

    if not headers:
        return [], []

    rows: list[dict[str, Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        value_cells = list(
            ws_values.iter_rows(
                min_row=row_idx, max_row=row_idx, min_col=min_col, max_col=min_col + len(headers) - 1, values_only=True
            )
        )[0]
        formula_cells = list(
            ws_formulas.iter_rows(
                min_row=row_idx, max_row=row_idx, min_col=min_col, max_col=min_col + len(headers) - 1, values_only=True
            )
        )[0]

        meaningful = False
        row: dict[str, Any] = {}
        for header, value, formula in zip(headers, value_cells, formula_cells):
            selected = clean_value(value)
            if selected is None and isinstance(formula, str) and not formula.startswith("="):
                selected = clean_value(formula)
            row[header] = selected

            formula_only = isinstance(formula, str) and formula.startswith("=") and selected in (None, "")
            if selected not in (None, "") and not formula_only:
                meaningful = True

        if meaningful:
            rows.append(row)

    return headers, rows


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: display(row.get(header)) for header in headers})


def markdown_table(headers: list[str], rows: list[dict[str, Any]], limit: int = 25) -> str:
    if not headers:
        return "_No data._\n"
    kept_headers = headers[: min(len(headers), 10)]
    lines = [
        "| " + " | ".join(md_escape(h) for h in kept_headers) + " |",
        "| " + " | ".join("---" for _ in kept_headers) + " |",
    ]
    for row in rows[:limit]:
        lines.append("| " + " | ".join(md_escape(row.get(h)) for h in kept_headers) + " |")
    if len(rows) > limit:
        lines.append(f"\n_Showing {limit} of {len(rows)} rows._")
    return "\n".join(lines) + "\n"


def write_sheet_markdown(sheet: str, headers: list[str], rows: list[dict[str, Any]], generated_at: str) -> Path:
    out = MD_DIR / f"{slug(sheet)}.md"
    csv_name = f"{slug(sheet)}.csv"
    content = f"""---
status: generated
type: tracker-view
source: "[[25_application_tracker_template.xlsx]]"
sheet: "{sheet}"
generated_at: "{generated_at}"
tags:
  - generated
  - tracker
---

# {sheet}

Source: [[25_application_tracker_template.xlsx]]

CSV: [[_obsidian/generated/csv/{csv_name}]]

Rows: {len(rows)}

{markdown_table(headers, rows)}
"""
    out.write_text(content, encoding="utf-8")
    return out


def count_by(rows: Iterable[dict[str, Any]], key: str) -> Counter[str]:
    c: Counter[str] = Counter()
    for row in rows:
        value = display(row.get(key)) or "Blank"
        c[value] += 1
    return c


def score_company(row: dict[str, Any]) -> float:
    try:
        fit = float(row.get("Fit") or 0)
        probability = float(row.get("Probability") or 0)
        compensation = float(row.get("Compensation") or 0)
        portfolio = float(row.get("Portfolio Fit") or 0)
    except (TypeError, ValueError):
        return 0.0
    return round((fit * 0.35) + (probability * 0.20) + (compensation * 0.20) + (portfolio * 0.25), 2)


def priority_number(row: dict[str, Any], key: str = "Priority") -> float:
    try:
        return float(row.get(key) or 999)
    except (TypeError, ValueError):
        return 999


def bullet_counts(title: str, counts: Counter[str]) -> str:
    lines = [f"## {title}", ""]
    if not counts:
        lines.append("_No data._")
    else:
        for key, value in counts.most_common():
            lines.append(f"- {key}: {value}")
    return "\n".join(lines)


def write_summary(all_data: dict[str, tuple[list[str], list[dict[str, Any]]]], generated_at: str) -> None:
    applications = all_data.get("Applications", ([], []))[1]
    companies = all_data.get("Companies", ([], []))[1]
    contacts = all_data.get("Contacts", ([], []))[1]
    assets = all_data.get("Content Assets", ([], []))[1]
    skills = all_data.get("Skill Sprint", ([], []))[1]

    top_companies = sorted(companies, key=score_company, reverse=True)[:15]
    open_assets = [row for row in assets if display(row.get("Status")).lower() not in {"done", "complete", "completed"}]
    open_skills = sorted(skills, key=priority_number)[:10]

    lines = [
        "---",
        "status: generated",
        "type: tracker-summary",
        'source: "[[25_application_tracker_template.xlsx]]"',
        f'generated_at: "{generated_at}"',
        "tags:",
        "  - generated",
        "  - tracker",
        "---",
        "",
        "# Tracker summary",
        "",
        "This file is generated from [[25_application_tracker_template.xlsx]]. Do not edit it manually.",
        "",
        f"Generated at: {generated_at}",
        "",
        "## Quick links",
        "",
    ]

    for sheet in SYNCED_SHEETS:
        lines.append(f"- [[{slug(sheet)}]]")

    lines.extend(
        [
            "",
            "## Totals",
            "",
            f"- Applications: {len(applications)}",
            f"- Companies: {len(companies)}",
            f"- Contacts: {len(contacts)}",
            f"- Open content assets: {len(open_assets)}",
            "",
            bullet_counts("Applications by status", count_by(applications, "Status")),
            "",
            bullet_counts("Companies by priority", count_by(companies, "Priority")),
            "",
            "## Top company targets by computed score",
            "",
        ]
    )

    if top_companies:
        headers = ["Company", "Priority", "Category", "Remote Likelihood", "Contractor Likelihood", "Status"]
        table_rows = []
        for row in top_companies:
            table_rows.append(
                {
                    "Company": row.get("Company"),
                    "Priority": row.get("Priority"),
                    "Category": row.get("Category"),
                    "Remote Likelihood": row.get("Remote Likelihood"),
                    "Contractor Likelihood": row.get("Contractor Likelihood"),
                    "Status": row.get("Status"),
                }
            )
        lines.append(markdown_table(headers, table_rows, limit=15))
    else:
        lines.append("_No company rows found._")

    lines.extend(["", "## Open content assets", ""])
    if open_assets:
        lines.append(markdown_table(["Asset", "Type", "Status", "Priority", "Due Date", "Dependency"], open_assets, limit=20))
    else:
        lines.append("_No open assets found._")

    lines.extend(["", "## Skill sprint", ""])
    if open_skills:
        lines.append(markdown_table(["Skill", "Priority", "Current Level", "Target Artifact", "Status", "Due Date"], open_skills, limit=10))
    else:
        lines.append("_No skill sprint rows found._")

    (MD_DIR / "tracker_summary.md").write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    CSV_DIR.mkdir(parents=True, exist_ok=True)
    MD_DIR.mkdir(parents=True, exist_ok=True)

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    wb_values = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=False)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    all_data: dict[str, tuple[list[str], list[dict[str, Any]]]] = {}
    for sheet in SYNCED_SHEETS:
        if sheet not in wb_formulas.sheetnames:
            continue
        headers, rows = extract_table(wb_values[sheet], wb_formulas[sheet])
        all_data[sheet] = (headers, rows)
        write_csv(CSV_DIR / f"{slug(sheet)}.csv", headers, rows)
        write_sheet_markdown(sheet, headers, rows, generated_at)

    write_summary(all_data, generated_at)

    print(f"Synced {len(all_data)} sheets from {WORKBOOK.name}")
    print(f"CSV: {CSV_DIR}")
    print(f"Markdown: {MD_DIR}")


if __name__ == "__main__":
    main()
